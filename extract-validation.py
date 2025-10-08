#!/usr/bin/env python3
"""
Extract Data Validation tooltips from Excel REPORT sheet
Outputs JSON mapping field IDs to validation messages
Usage: python3 extract-validation.py <path-to-excel-file>
"""

import sys
import json
from openpyxl import load_workbook

# Import the cell mapping from ExcelMapper.js
# These match the excelReportInputMapping in 4011-ExcelMapper.js
CELL_TO_FIELD_MAPPING = {
    # Section 02: Building Information
    'D12': 'd_12',  # Major Occupancy
    'D13': 'd_13',  # Reference Standard
    'D14': 'd_14',  # Actual/Target Use
    'D15': 'd_15',  # Carbon Standard
    'H12': 'h_12',  # Reporting Period
    'H13': 'h_13',  # Service Life
    'H14': 'h_14',  # Project Name
    'H15': 'h_15',  # Conditioned Area
    'H16': 'i_16',  # Certifier
    'H17': 'i_17',  # License No
    'L12': 'l_12',  # Elec Cost
    'L13': 'l_13',  # Gas Cost
    'L14': 'l_14',  # Propane Cost
    'L15': 'l_15',  # Wood Cost
    'L16': 'l_16',  # Oil Cost

    # Section 03: Climate
    'D19': 'd_19',  # Province
    'H19': 'h_19',  # City
    'H20': 'h_20',  # Present/Future Weather
    'H21': 'h_21',  # Capacitance/Static
    'I21': 'i_21',  # Capacitance Slider
    'M19': 'm_19',  # Days Cooling

    # Section 04: Actual vs Target Energy
    'D27': 'd_27',  # Actual Elec Use
    'D28': 'd_28',  # Actual Gas Use
    'D29': 'd_29',  # Actual Propane Use
    'D30': 'd_30',  # Actual Oil Use
    'D31': 'd_31',  # Actual Wood Use
    'L27': 'l_27',  # Elec Emission Factor
    'L28': 'l_28',  # Gas Emission Factor
    'L29': 'l_29',  # Propane Emission Factor
    'L30': 'l_30',  # Oil Emission Factor
    'L31': 'l_31',  # Wood Emission Factor
    'H35': 'h_35',  # PER Factor

    # Section 05: Emissions
    'D39': 'd_39',  # Construction Type
    'I41': 'i_41',  # Modelled Embodied Carbon

    # Section 06: Renewable Energy
    'D44': 'd_44',  # Photovoltaics
    'D45': 'd_45',  # Wind
    'D46': 'd_46',  # Remove EV Charging
    'I45': 'i_45',  # WWS Electricity
    'K45': 'k_45',  # Green Natural Gas
    'I46': 'i_46',  # Other Removals
    'M43': 'm_43',  # P.5 Exterior/Site/Other Loads

    # Section 07: Water Use
    'D49': 'd_49',  # Water Use Method
    'E49': 'e_49',  # User Defined Water Use
    'E50': 'e_50',  # By Engineer DHW
    'D51': 'd_51',  # DHW Energy Source
    'D52': 'd_52',  # DHW EF/COP
    'D53': 'd_53',  # DHW Recovery Eff
    'K52': 'k_52',  # SHW AFUE
    'D54': 'd_54',  # Total WWR %

    # Add more as needed - see 4011-ExcelMapper.js lines 40-270 for full mapping
}


def extract_validation_tooltips(excel_path, scan_mode='mapped'):
    """Extract data validation input messages from REPORT sheet

    Args:
        excel_path: Path to Excel file
        scan_mode: 'mapped' (use predefined mapping) or 'comprehensive' (scan rows 12-145)
    """

    try:
        from openpyxl.utils import get_column_letter

        # Load workbook with data_only=False to preserve validation
        wb = load_workbook(excel_path, data_only=False)

        if 'REPORT' not in wb.sheetnames:
            print(f"ERROR: REPORT sheet not found in {excel_path}", file=sys.stderr)
            print(f"Available sheets: {wb.sheetnames}", file=sys.stderr)
            return None

        ws = wb['REPORT']
        tooltips = {}
        validation_count = 0

        if scan_mode == 'comprehensive':
            # Comprehensive scan: rows 12-145, columns A-P (1-16)
            print(f"🔍 Scanning rows 12-145, columns A-P comprehensively...", file=sys.stderr)

            # Reverse mapping: cell_ref -> field_id
            cell_to_field = {v: k for k, v in CELL_TO_FIELD_MAPPING.items()}

            for row in range(12, 146):  # Rows 12-145 inclusive
                for col in range(1, 17):  # Columns A-P (1-16)
                    cell_ref = f"{get_column_letter(col)}{row}"
                    cell = ws[cell_ref]

                    # Check if cell has data validation
                    if hasattr(ws, 'data_validations'):
                        for dv in ws.data_validations.dataValidation:
                            if cell.coordinate in dv.cells:
                                if dv.promptTitle or dv.prompt:
                                    # Try to find field_id from mapping, otherwise use cell_ref
                                    field_id = cell_to_field.get(cell_ref, cell_ref)

                                    tooltips[field_id] = {
                                        'cell': cell_ref,
                                        'title': dv.promptTitle or '',
                                        'message': dv.prompt or '',
                                    }
                                    validation_count += 1
                                    break
        else:
            # Mapped mode: use predefined cell mapping
            print(f"🔍 Scanning predefined mapped cells...", file=sys.stderr)

            for cell_ref, field_id in CELL_TO_FIELD_MAPPING.items():
                cell = ws[cell_ref]

                # Check if cell has data validation
                if hasattr(ws, 'data_validations'):
                    for dv in ws.data_validations.dataValidation:
                        # Check if this validation applies to our cell
                        if cell.coordinate in dv.cells:
                            if dv.promptTitle or dv.prompt:
                                tooltips[field_id] = {
                                    'cell': cell_ref,
                                    'title': dv.promptTitle or '',
                                    'message': dv.prompt or '',
                                }
                                validation_count += 1
                                break

        print(f"✅ Extracted {validation_count} validation tooltips from REPORT sheet", file=sys.stderr)
        return tooltips

    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return None


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 extract-validation.py <path-to-excel-file> [--comprehensive]")
        print("\nModes:")
        print("  Default: Extract only predefined mapped cells")
        print("  --comprehensive: Scan all cells in rows 12-145, columns A-P")
        print("\nExamples:")
        print('  python3 extract-validation.py "OBJECTIVE 4011RF/TEUI-4011RF.xlsx"')
        print('  python3 extract-validation.py "OBJECTIVE 4011RF/TEUI-4011RF.xlsx" --comprehensive')
        sys.exit(1)

    excel_path = sys.argv[1]
    scan_mode = 'comprehensive' if '--comprehensive' in sys.argv else 'mapped'

    tooltips = extract_validation_tooltips(excel_path, scan_mode=scan_mode)

    if tooltips is not None:
        # Output JSON to stdout
        print(json.dumps(tooltips, indent=2))

        # Optionally save to file
        output_file = 'OBJECTIVE 4011RF/data/validation-tooltips.json'
        try:
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(tooltips, f, indent=2)
            print(f"✅ Saved to {output_file}", file=sys.stderr)
        except Exception as e:
            print(f"⚠️  Could not save to file: {e}", file=sys.stderr)


if __name__ == '__main__':
    main()
