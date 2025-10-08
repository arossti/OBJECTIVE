#!/usr/bin/env python3
"""
Extract ALL Data Validation tooltips from rows 12-17 of REPORT sheet (Section 02)
This scans all columns to find any cells with validation, not just known mappings
Usage: python3 extract-validation-section02.py <path-to-excel-file>
"""

import sys
import json
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def extract_section02_tooltips(excel_path):
    """Extract all data validation from rows 12-17 (Section 02)"""

    try:
        wb = load_workbook(excel_path, data_only=False)

        if 'REPORT' not in wb.sheetnames:
            print(f"ERROR: REPORT sheet not found", file=sys.stderr)
            return None

        ws = wb['REPORT']
        tooltips = {}
        validation_count = 0

        # Scan rows 12-17, columns A through P (covering all likely fields)
        for row in range(12, 18):  # Rows 12-17 inclusive
            for col in range(1, 17):  # Columns A-P (1-16)
                cell_ref = f"{get_column_letter(col)}{row}"
                cell = ws[cell_ref]

                # Check if this cell has data validation
                if hasattr(ws, 'data_validations'):
                    for dv in ws.data_validations.dataValidation:
                        if cell.coordinate in dv.cells:
                            if dv.promptTitle or dv.prompt:
                                tooltips[cell_ref] = {
                                    'title': dv.promptTitle or '',
                                    'message': dv.prompt or '',
                                    'row': row,
                                    'col': get_column_letter(col)
                                }
                                validation_count += 1
                                print(f"  Found: {cell_ref} - {dv.promptTitle}", file=sys.stderr)
                                break

        print(f"\n✅ Found {validation_count} validation tooltips in rows 12-17", file=sys.stderr)
        return tooltips

    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc(file=sys.stderr)
        return None


def main():
    if len(sys.argv) != 2:
        print("Usage: python3 extract-validation-section02.py <path-to-excel-file>")
        sys.exit(1)

    excel_path = sys.argv[1]
    tooltips = extract_section02_tooltips(excel_path)

    if tooltips is not None:
        # Output JSON to stdout
        print(json.dumps(tooltips, indent=2))

        # Save to file
        output_file = 'OBJECTIVE 4011RF/data/section02-validation-tooltips.json'
        try:
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(tooltips, f, indent=2)
            print(f"✅ Saved to {output_file}", file=sys.stderr)
        except Exception as e:
            print(f"⚠️  Could not save: {e}", file=sys.stderr)


if __name__ == '__main__':
    main()
